import openpyxl as px
from Member import Member


class Extractor:
    def __init__(self, workBookPath):
        # load workbook and sheet
        self.workbook = px.load_workbook(workBookPath)
        self.sheet = self.workbook.get_sheet_by_name(
            self.workbook.get_sheet_names()[0]
        )

        """
             max rows in sheet
             can also include empty cells, so condition should be
             there
        """
        self.maxRows = self.sheet.max_row

    def extractAllEmailAddress(self):
        emailAddresses = []
        for i in range(2, self.maxRows + 1):
            emailAddress = self.sheet.cell(row=i, column=4).value

            if emailAddress != None:
                emailAddresses.append(emailAddress)

        return emailAddresses

    def extractRecordByTeam(self, teamName):
        teamData = []
        teamColumn = 3

        for i in range(2, self.maxRows + 1):
            # check for team name match
            if self.sheet.cell(row=i, column=teamColumn).value == teamName:
                name = self.sheet.cell(row=i, column=1).value
                nsuID = self.sheet.cell(row=i, column=2).value
                email = self.sheet.cell(row=i, column=4).value

                if email == None:
                    email = "N/A"

                # create a Member object
                member = Member(name=name, nsuId=nsuID, team=teamName, email=email)
                teamData.append(member)

        return teamData

    def extractAllTeamRecords(self):
        teamDict = {}
        teamNames = ["Corporate", "Operations", "Publications", "Promotions", "Logistics"]
        teamColumn = 3

        for team in teamNames:
            teamData = []
            for i in range(2, self.maxRows + 1):
                # check for team name match
                if self.sheet.cell(row=i, column=teamColumn).value == team:
                    name = self.sheet.cell(row=i, column=1).value
                    nsuID = self.sheet.cell(row=i, column=2).value
                    email = self.sheet.cell(row=i, column=4).value

                    # create a Member object
                    member = Member(name=name, nsuId=nsuID, team=team, email=email)
                    teamData.append(member)

            teamDict[team] = teamData

        return teamDict

    def getEmailByTeam(self, teamName):
        teamMail = []
        teamColumn = 3

        for i in range(2, self.maxRows + 1):
            # check for team name match
            if self.sheet.cell(row=i, column=teamColumn).value == teamName:
                email = self.sheet.cell(row=i, column=4).value
                teamMail.append(email)

        return teamMail
